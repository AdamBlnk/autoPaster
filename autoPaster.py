from openpyxl import load_workbook

wb = load_workbook("excelExample.xlsx")
mySheet = wb.worksheets[0]
currentString = ""

# Check the 19th column, if there is a value there, assign it to 'currentString'
# If the column is empty, assign that cell to the value of 'currentString'
for i in range(1, mySheet.max_row + 1, 1):
    if mySheet.cell(i,19).value is not None:
        currentString = mySheet.cell(i,19).value
    else:
        mySheet.cell(i,19).value = currentString
    
wb.save("excelExample.xlsx")